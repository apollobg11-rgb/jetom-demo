#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ЖЕТОМ ТРАНС — Калкулатор за командировъчни
Фази 2-3: GPS2 парсване + Mapping + Групиране по шофьори
"""

from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
import openpyxl
from collections import defaultdict
import traceback
import csv
import io
import tempfile
import os

app = Flask(__name__)

# Константи
EUR_TO_BGN = 1.95583
DEFAULT_EUR_RATE = {
    'Гърция': 43,
    'Румъния': 46,
    'Турция': 43,
    'Чужбина (неопределена)': 43
}


def detect_country(address):
    """Детекция на държава от адрес"""
    if not address:
        return "Неизвестна"
    
    addr = str(address)
    
    # България — на кирилица
    if "България" in addr:
        return "България"
    
    # Гърция — на гръцки
    greek_markers = ["Δήμος", "Περιφερ", "Ελληνικ", "Δημοτικ", "Κοινότητα"]
    if any(m in addr for m in greek_markers):
        return "Гърция"
    
    # Румъния
    if "Румъния" in addr or "România" in addr or "Romania" in addr:
        return "Румъния"
    
    # Турция
    if "Турция" in addr or "Türkiye" in addr or "Turkey" in addr:
        return "Турция"
    
    return "Чужбина (неопределена)"


def parse_gps1(file_path):
    """Парсва GPS Система 1 файл (.xlsx)"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    records = []
    
    for row in ws.iter_rows(min_row=9, values_only=True):
        truck = row[0]
        
        if not truck or truck == "Общо":
            continue
            
        start_time = row[1]
        end_time = row[2]
        from_addr = row[3]
        to_addr = row[4]
        
        # Конвертираме датите
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
        
        from_country = detect_country(from_addr)
        to_country = detect_country(to_addr)
        
        records.append({
            'truck': truck,
            'start_time': start_time,
            'end_time': end_time,
            'from_addr': from_addr,
            'to_addr': to_addr,
            'from_country': from_country,
            'to_country': to_country,
            'source': 'GPS1'
        })
    
    return records


def parse_gps2(file_path):
    """
    Парсва GPS Система 2 файл (.xlsx)
    Множество sheets, всеки sheet = 1 камион
    """
    wb = openpyxl.load_workbook(file_path)
    records = []
    
    for sheet_name in wb.sheetnames:
        # Sheet name = регистрационен номер на камиона
        truck = sheet_name.strip()
        ws = wb[sheet_name]
        
        # Header е на ред 1, данните започват от ред 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[4]:  # Колона E: Начална дата
                continue
            
            start_addr = row[2]  # Колона C: Начален адрес
            start_time = row[4]  # Колона E: Начална дата
            end_addr = row[10]   # Колона K: Краен адрес
            end_time = row[12]   # Колона M: Крайна дата
            
            # Парсваме датите (формат: DD/MM/YYYY HH:MM:SS)
            if isinstance(start_time, str):
                try:
                    start_time = datetime.strptime(start_time, '%d/%m/%Y %H:%M:%S')
                except:
                    try:
                        start_time = datetime.strptime(start_time, '%d/%m/%Y %H:%M')
                    except:
                        continue
            
            if isinstance(end_time, str):
                try:
                    end_time = datetime.strptime(end_time, '%d/%m/%Y %H:%M:%S')
                except:
                    try:
                        end_time = datetime.strptime(end_time, '%d/%m/%Y %H:%M')
                    except:
                        continue
            
            from_country = detect_country(start_addr)
            to_country = detect_country(end_addr)
            
            records.append({
                'truck': truck,
                'start_time': start_time,
                'end_time': end_time,
                'from_addr': start_addr,
                'to_addr': end_addr,
                'from_country': from_country,
                'to_country': to_country,
                'source': 'GPS2'
            })
    
    return records


def parse_mapping(file_path):
    """
    Парсва mapping CSV (Камион;Шофьор)
    Returns: dict { truck: driver_name }
    """
    mapping = {}
    
    # Опитваме се да отворим като CSV
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Детектираме delimiter (';' или ',')
        delimiter = ';' if ';' in content.split('\n')[0] else ','
        
        lines = content.strip().split('\n')
        for i, line in enumerate(lines):
            if i == 0:  # Skip header
                continue
            
            parts = line.split(delimiter)
            if len(parts) >= 2:
                truck = parts[0].strip()
                driver = parts[1].strip()
                mapping[truck] = driver
    
    except Exception as e:
        print(f"Error parsing mapping: {e}")
    
    return mapping


def build_travel_blocks(records):
    """Построява travel blocks от GPS записи"""
    by_truck = defaultdict(list)
    for rec in records:
        by_truck[rec['truck']].append(rec)
    
    all_blocks = {}
    
    for truck, truck_records in by_truck.items():
        truck_records.sort(key=lambda x: x['start_time'])
        
        blocks = []
        current_block = None
        
        for rec in truck_records:
            from_country = rec['from_country']
            to_country = rec['to_country']
            
            # България → Чужбина = START
            if from_country == "България" and to_country != "България":
                if current_block:
                    blocks.append(current_block)
                
                current_block = {
                    'start_date': rec['start_time'],
                    'end_date': rec['end_time'],
                    'country': to_country,
                    'records': [rec]
                }
            
            # Чужбина → Чужбина = CONTINUE
            elif from_country != "България" and to_country != "България":
                if current_block:
                    current_block['end_date'] = rec['end_time']
                    current_block['country'] = to_country
                    current_block['records'].append(rec)
                else:
                    current_block = {
                        'start_date': rec['start_time'],
                        'end_date': rec['end_time'],
                        'country': to_country,
                        'records': [rec]
                    }
            
            # Чужбина → България = END
            elif from_country != "България" and to_country == "България":
                if current_block:
                    current_block['end_date'] = rec['start_time']
                    current_block['records'].append(rec)
                    blocks.append(current_block)
                    current_block = None
        
        if current_block:
            blocks.append(current_block)
        
        # Изчисляваме дни
        for block in blocks:
            start = block['start_date']
            end = block['end_date']
            
            if isinstance(start, datetime) and isinstance(end, datetime):
                start_date = start.date()
                end_date = end.date()
                days = (end_date - start_date).days + 1
                block['days'] = max(1, days)
            else:
                block['days'] = 1
            
            if isinstance(start, datetime):
                block['start_date_formatted'] = start.strftime('%d.%m.%Y')
            if isinstance(end, datetime):
                block['end_date_formatted'] = end.strftime('%d.%m.%Y')
        
        all_blocks[truck] = blocks
    
    return all_blocks


def group_by_driver(blocks, mapping):
    """
    Групира travel blocks по шофьор
    Returns: dict { driver: [blocks] }
    """
    by_driver = defaultdict(list)
    unmapped_trucks = []
    
    for truck, truck_blocks in blocks.items():
        driver = mapping.get(truck)
        
        if driver:
            for block in truck_blocks:
                block_copy = block.copy()
                block_copy['truck'] = truck
                by_driver[driver].append(block_copy)
        else:
            unmapped_trucks.append(truck)
            # Добавяме към "Неразпределени"
            for block in truck_blocks:
                block_copy = block.copy()
                block_copy['truck'] = truck
                by_driver['⚠️ Неразпределени'].append(block_copy)
    
    # Сортираме блоковете по дата за всеки шофьор
    for driver in by_driver:
        by_driver[driver].sort(key=lambda x: x['start_date'])
    
    return dict(by_driver), unmapped_trucks


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process_files():
    """Обработка на качените файлове"""
    try:
        # Проверка за файлове
        if 'gps1_file' not in request.files:
            return jsonify({'error': 'Моля качете GPS Система 1 файл'}), 400
        
        gps1_file = request.files['gps1_file']
        gps2_file = request.files.get('gps2_file')
        mapping_file = request.files.get('mapping_file')
        
        if gps1_file.filename == '':
            return jsonify({'error': 'Моля качете GPS Система 1 файл'}), 400
        
        # Записваме временно файловете
        temp_dir = tempfile.mkdtemp()
        
        gps1_path = os.path.join(temp_dir, 'gps1.xlsx')
        gps1_file.save(gps1_path)
        
        # Парсваме GPS1
        records = parse_gps1(gps1_path)
        
        # Парсваме GPS2 ако е качен
        if gps2_file and gps2_file.filename:
            gps2_path = os.path.join(temp_dir, 'gps2.xlsx')
            gps2_file.save(gps2_path)
            records.extend(parse_gps2(gps2_path))
        
        # Парсваме mapping ако е качен
        mapping = {}
        if mapping_file and mapping_file.filename:
            mapping_path = os.path.join(temp_dir, 'mapping.csv')
            mapping_file.save(mapping_path)
            mapping = parse_mapping(mapping_path)
        
        # Строим travel blocks
        blocks = build_travel_blocks(records)
        
        # Ако има mapping, групираме по шофьор
        if mapping:
            by_driver, unmapped = group_by_driver(blocks, mapping)
            result_data = []
            
            for driver in sorted(by_driver.keys()):
                driver_blocks = by_driver[driver]
                
                for i, block in enumerate(driver_blocks, 1):
                    result_data.append({
                        'driver': driver,
                        'truck': block['truck'],
                        'block_num': i,
                        'start_date': block['start_date'].strftime('%d.%m.%Y %H:%M') if isinstance(block['start_date'], datetime) else str(block['start_date']),
                        'end_date': block['end_date'].strftime('%d.%m.%Y %H:%M') if isinstance(block['end_date'], datetime) else str(block['end_date']),
                        'country': block['country'],
                        'days': block['days'],
                        'eur_rate': DEFAULT_EUR_RATE.get(block['country'], 43),
                        'records_count': len(block['records'])
                    })
            
            # Статистика
            total_drivers = len([d for d in by_driver.keys() if d != '⚠️ Неразпределени'])
            total_blocks_mapped = sum(len(b) for d, b in by_driver.items() if d != '⚠️ Неразпределени')
            
            stats = {
                'total_records': len(records),
                'abroad_records': sum(1 for r in records if r['to_country'] != 'България'),
                'abroad_percentage': round(sum(1 for r in records if r['to_country'] != 'България') / len(records) * 100, 1) if len(records) > 0 else 0,
                'total_trucks': len(blocks),
                'trucks_with_travel': len([t for t, b in blocks.items() if len(b) > 0]),
                'total_blocks': sum(len(b) for b in blocks.values()),
                'total_drivers': total_drivers,
                'unmapped_trucks': unmapped,
                'has_mapping': True
            }
        
        else:
            # Без mapping - показваме по камиони
            result_data = []
            for truck in sorted(blocks.keys()):
                truck_blocks = blocks[truck]
                if len(truck_blocks) == 0:
                    continue
                
                for i, block in enumerate(truck_blocks, 1):
                    result_data.append({
                        'driver': None,
                        'truck': truck,
                        'block_num': i,
                        'start_date': block['start_date'].strftime('%d.%m.%Y %H:%M') if isinstance(block['start_date'], datetime) else str(block['start_date']),
                        'end_date': block['end_date'].strftime('%d.%m.%Y %H:%M') if isinstance(block['end_date'], datetime) else str(block['end_date']),
                        'country': block['country'],
                        'days': block['days'],
                        'eur_rate': DEFAULT_EUR_RATE.get(block['country'], 43),
                        'records_count': len(block['records'])
                    })
            
            stats = {
                'total_records': len(records),
                'abroad_records': sum(1 for r in records if r['to_country'] != 'България'),
                'abroad_percentage': round(sum(1 for r in records if r['to_country'] != 'България') / len(records) * 100, 1) if len(records) > 0 else 0,
                'total_trucks': len(blocks),
                'trucks_with_travel': len([t for t, b in blocks.items() if len(b) > 0]),
                'total_blocks': sum(len(b) for b in blocks.values()),
                'has_mapping': False
            }
        
        # Почистваме временните файлове
        import shutil
        shutil.rmtree(temp_dir)
        
        return jsonify({
            'success': True,
            'stats': stats,
            'data': result_data
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при обработка: {str(e)}'}), 500


@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Генерира Excel файл с командировките"""
    try:
        data = request.json.get('data', [])
        has_mapping = request.json.get('has_mapping', False)
        
        # Създаваме workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Командировки"
        
        # Header
        if has_mapping:
            headers = ['№', 'Шофьор', 'Камион', 'Дата от', 'Дата до', 'Дни', 'EUR/ден', 'Сума EUR', 'Сума BGN']
        else:
            headers = ['№', 'Камион', 'Дата от', 'Дата до', 'Дни', 'EUR/ден', 'Сума EUR', 'Сума BGN']
        
        # Стилизация на header
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Пишем header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Данни
        row_num = 2
        current_driver = None
        driver_start_row = 2
        row_counter = 1
        
        driver_totals = defaultdict(lambda: {'days': 0, 'eur': 0, 'bgn': 0, 'blocks': 0})
        
        for block in data:
            driver = block.get('driver') or block.get('truck')
            truck = block.get('truck', '')
            start_date = block.get('start_date', '').split(' ')[0]  # Само датата
            end_date = block.get('end_date', '').split(' ')[0]
            days = block.get('days', 0)
            eur_rate = block.get('eur_rate', 43)
            eur_sum = days * eur_rate
            bgn_sum = eur_sum * EUR_TO_BGN
            
            # Ако сменяме шофьор, добавяме подсума
            if current_driver and driver != current_driver:
                # Подсума за предишния шофьор
                ws.merge_cells(f'A{row_num}:B{row_num}' if has_mapping else f'A{row_num}:A{row_num}')
                subtotal_cell = ws.cell(row=row_num, column=1, value='За получаване')
                subtotal_cell.font = Font(bold=True, size=11)
                subtotal_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                col_offset = 3 if has_mapping else 2
                ws.cell(row=row_num, column=col_offset+2, value=driver_totals[current_driver]['days']).font = Font(bold=True)
                ws.cell(row=row_num, column=col_offset+4, value=driver_totals[current_driver]['eur']).font = Font(bold=True)
                ws.cell(row=row_num, column=col_offset+5, value=driver_totals[current_driver]['bgn']).font = Font(bold=True)
                
                # Apply styling
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    cell.border = border
                
                row_num += 1
                row_counter = 1
            
            # Данни за реда
            if has_mapping:
                values = [row_counter, driver if driver != current_driver else '', truck, start_date, end_date, days, eur_rate, eur_sum, bgn_sum]
            else:
                values = [row_counter, truck if driver != current_driver else '', start_date, end_date, days, eur_rate, eur_sum, bgn_sum]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                
                # Bold за име на шофьор/камион
                if col == 2 and value:
                    cell.font = Font(bold=True, size=11)
                
                # Alignment
                if col <= (3 if has_mapping else 2):
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Format numbers
                if col >= len(headers) - 1:  # EUR и BGN колони
                    cell.number_format = '#,##0.00'
            
            # Accumulate totals
            driver_totals[driver]['days'] += days
            driver_totals[driver]['eur'] += eur_sum
            driver_totals[driver]['bgn'] += bgn_sum
            driver_totals[driver]['blocks'] += 1
            
            current_driver = driver
            row_num += 1
            row_counter += 1
        
        # Последна подсума
        if current_driver:
            ws.merge_cells(f'A{row_num}:B{row_num}' if has_mapping else f'A{row_num}:A{row_num}')
            subtotal_cell = ws.cell(row=row_num, column=1, value='За получаване')
            subtotal_cell.font = Font(bold=True, size=11)
            subtotal_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            col_offset = 3 if has_mapping else 2
            ws.cell(row=row_num, column=col_offset+2, value=driver_totals[current_driver]['days']).font = Font(bold=True)
            ws.cell(row=row_num, column=col_offset+4, value=driver_totals[current_driver]['eur']).font = Font(bold=True)
            ws.cell(row=row_num, column=col_offset+5, value=driver_totals[current_driver]['bgn']).font = Font(bold=True)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        if has_mapping:
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
        else:
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Генерираме име на файл с дата
        from datetime import datetime as dt
        filename = f"Командировки_{dt.now().strftime('%Y_%m')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при генериране на Excel: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
